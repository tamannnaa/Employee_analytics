import os
import uuid
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows

# Generate an Excel Report
def generate_employee_report(employees, filename=None):
    if not employees:
        raise ValueError("No employee data provided.")

    os.makedirs("exports", exist_ok=True)

    if not filename:
        filename = f"employee_report_{uuid.uuid4().hex[:6]}.xlsx"

    filepath = os.path.join("exports", filename)

    df = pd.DataFrame(employees)
    df.drop(columns=["_id"], inplace=True, errors='ignore')

    for col in df.columns:
        if df[col].apply(lambda x: isinstance(x, list)).any():
            df[col] = df[col].apply(lambda x: ", ".join(map(str, x)) if isinstance(x, list) else x)

    # Department Summary
    dept_summary = (
        df.groupby("department")
        .agg(employee_count=("employee_id", "count"), avg_salary=("salary", "mean"))
        .reset_index()
    )

    # Salary Range Analysis
    salary_bins = [0, 50_000, 75_000, 100_000, 150_000, 200_000]
    df["salary_range"] = pd.cut(df["salary"], bins=salary_bins)
    
    salary_analysis = (
        df["salary_range"]
        .value_counts()
        .sort_index()
        .reset_index()
    )
    salary_analysis.columns = ["salary_range", "employee_count"]
    salary_analysis["salary_range"] = salary_analysis["salary_range"].astype(str)


    # Excel Workbook Creation
    wb = Workbook()

    # --- Sheet 1: Employee List ---
    ws1 = wb.active
    ws1.title = "Employee List"

    for row in dataframe_to_rows(df.drop(columns=["salary_range"]), index=False, header=True):
        ws1.append(row)

    style_worksheet(ws1)

    # Highlight top performers
    performance_col = df.columns.get_loc("performance_score") + 1
    for row in range(2, ws1.max_row + 1):
        performance = ws1.cell(row=row, column=performance_col).value
        if performance is not None and performance >= 4.5:
            for col in range(1, ws1.max_column + 1):
                ws1.cell(row=row, column=col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # --- Sheet 2: Department Summary ---
    ws2 = wb.create_sheet("Department Summary")

    for row in dataframe_to_rows(dept_summary, index=False, header=True):
        ws2.append(row)

    style_worksheet(ws2)
    title_cell = f"B{ws2.max_row + 2}"
    ws2[title_cell] = "Employees by Department"
    ws2[title_cell].font = Font(bold=True, size=14)

    # Add 2 empty rows after title
    ws2.insert_rows(ws2.max_row + 1, amount=2)

    # Bar Chart - Employees by Department
    chart = BarChart()
    chart.height = 10  # Controls vertical size of the chart

    data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row, max_col=2)
    categories = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws2.add_chart(chart, "B10")  # Positioned lower to avoid title overlap

    # --- Sheet 3: Salary Analysis ---
    ws3 = wb.create_sheet("Salary Analysis")

    for row in dataframe_to_rows(salary_analysis, index=False, header=True):
        ws3.append(row)

    style_worksheet(ws3)
    title_cell = f"B{ws3.max_row + 2}"
    ws3[title_cell] = "Salary Distribution"
    ws3[title_cell].font = Font(bold=True, size=14)

    # Add 2 empty rows after title
    ws3.insert_rows(ws3.max_row + 1, amount=2)

    # Pie Chart - Salary Distribution
    pie = PieChart()

    data = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row)
    labels = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws3.add_chart(pie, "B10")

    # --------------------------------------------
    # Save the Workbook
    # --------------------------------------------
    wb.save(filepath)
    return filepath


def style_worksheet(ws):

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue header
    header_font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Style headers
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

        # Auto-adjust column widths
        col_letter = get_column_letter(col_num)
        max_length = max(len(str(cell.value)) for cell in ws[col_letter])
        ws.column_dimensions[col_letter].width = max(max_length + 2, 15)

    # Style data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment
            cell.border = thin_border
