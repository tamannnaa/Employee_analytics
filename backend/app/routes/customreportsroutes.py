from fastapi import APIRouter, HTTPException, BackgroundTasks, Query
from typing import List, Dict, Any, Optional, Union
from pydantic import BaseModel
import uuid
import os
import pandas as pd
from datetime import datetime
from pathlib import Path
from io import BytesIO
from ..database import get_database
from openpyxl.styles import Font
from datetime import datetime
# from ..excel_generator import generate_custom_report

router = APIRouter(prefix="/reports/custom", tags=["Custom Report Builder"])
db = get_database()
employees_collection = db["employees"]
custom_reports_collection = db["custom_reports"]
custom_templates_collection = db["custom_report_templates"]

# Ensure custom reports directory exists
CUSTOM_REPORTS_DIRECTORY = Path("custom_reports")
CUSTOM_REPORTS_DIRECTORY.mkdir(exist_ok=True)

# Pydantic Models
class FilterCondition(BaseModel):
    field: str
    operator: str
    value: str

class CustomReportConfig(BaseModel):
    name: str
    description: str = ""
    selectedFields: List[str]
    filters: List[FilterCondition] = []
    sortBy: str = "name"
    sortOrder: str = "asc"  # "asc" or "desc"
    includeCharts: bool = True
    format: str = "xlsx"  # "xlsx" or "csv"

class PreviewRequest(BaseModel):
    selectedFields: List[str]
    filters: List[FilterCondition] = []
    sortBy: str = "name"
    sortOrder: str = "asc"
    limit: int = 10

class SaveTemplateRequest(BaseModel):
    name: str
    description: str = ""
    config: CustomReportConfig

def build_mongodb_query(filters: List[FilterCondition]) -> Dict[str, Any]:
    """Convert filter conditions to MongoDB query"""
    if not filters:
        return {}
    
    query_conditions = []
    
    for filter_cond in filters:
        field = filter_cond.field
        operator = filter_cond.operator
        value = filter_cond.value
        
        if not value.strip():  # Skip empty values
            continue
            
        try:
            if operator == "equals":
                if field in ["salary", "performance_score"]:
                    query_conditions.append({field: float(value)})
                elif field == "is_active":
                    query_conditions.append({field: value.lower() in ["true", "1", "yes"]})
                else:
                    query_conditions.append({field: {"$regex": f"^{value}$", "$options": "i"}})
                    
            elif operator == "contains":
                query_conditions.append({field: {"$regex": value, "$options": "i"}})
                
            elif operator == "greater_than":
                if field in ["salary", "performance_score"]:
                    query_conditions.append({field: {"$gt": float(value)}})
                elif field in ["join_date"]:
                    query_conditions.append({field: {"$gt": value}})
                    
            elif operator == "less_than":
                if field in ["salary", "performance_score"]:
                    query_conditions.append({field: {"$lt": float(value)}})
                elif field in ["join_date"]:
                    query_conditions.append({field: {"$lt": value}})
                    
            elif operator == "between":
                # Expecting format: "min,max"
                values = [v.strip() for v in value.split(",")]
                if len(values) == 2:
                    if field in ["salary", "performance_score"]:
                        query_conditions.append({field: {"$gte": float(values[0]), "$lte": float(values[1])}})
                    elif field in ["join_date"]:
                        query_conditions.append({field: {"$gte": values[0], "$lte": values[1]}})
                        
            elif operator == "in":
                # Expecting comma-separated values
                values = [v.strip() for v in value.split(",")]
                query_conditions.append({field: {"$in": values}})
                
        except (ValueError, TypeError) as e:
            # Skip invalid filter conditions
            continue
    
    if query_conditions:
        return {"$and": query_conditions} if len(query_conditions) > 1 else query_conditions[0]
    
    return {}

def get_sort_direction(sort_order: str) -> int:
    """Convert sort order string to MongoDB sort direction"""
    return 1 if sort_order.lower() == "asc" else -1

@router.post("/preview")
async def preview_custom_report(request: PreviewRequest):
    """Preview data for custom report configuration"""
    try:
        # Build MongoDB query from filters
        mongo_query = build_mongodb_query(request.filters)
        
        # Create projection for selected fields
        projection = {field: 1 for field in request.selectedFields}
        projection["_id"] = 0  # Exclude MongoDB _id
        
        # Get sort direction
        sort_direction = get_sort_direction(request.sortOrder)
        
        # Execute query with limit
        cursor = employees_collection.find(
            mongo_query, 
            projection
        ).sort(request.sortBy, sort_direction).limit(request.limit)
        
        employees = list(cursor)
        
        # Process data for preview
        for employee in employees:
            for field, value in employee.items():
                if isinstance(value, list):
                    employee[field] = ", ".join(map(str, value))
        
        return employees
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Preview failed: {str(e)}")

def _generate_custom_report_task(report_id: str, config: CustomReportConfig):
    """Background task to generate custom report"""
    try:
        print(f"Starting report generation for {report_id}")
        print(f"Custom reports directory: {CUSTOM_REPORTS_DIRECTORY.absolute()}")
        
        # Update status to generating
        custom_reports_collection.update_one(
            {"_id": report_id},
            {"$set": {
                "status": "Generating",
                "updated_at": datetime.utcnow().isoformat()
            }}
        )

        # Build MongoDB query from filters
        mongo_query = build_mongodb_query(config.filters)
        
        # Create projection for selected fields
        projection = {field: 1 for field in config.selectedFields}
        projection["_id"] = 0  # Exclude MongoDB _id
        
        # Get sort direction
        sort_direction = get_sort_direction(config.sortOrder)
        
        # Execute query
        cursor = employees_collection.find(
            mongo_query, 
            projection
        ).sort(config.sortBy, sort_direction)
        
        employees = list(cursor)

        if not employees:
            custom_reports_collection.update_one(
                {"_id": report_id},
                {"$set": {
                    "status": "Failed", 
                    "message": "No employees match the specified criteria.",
                    "updated_at": datetime.utcnow().isoformat()
                }}
            )
            return

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(c for c in config.name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        filename = f"{safe_name.replace(' ', '_')}_{timestamp}.{config.format}"
        filepath = CUSTOM_REPORTS_DIRECTORY / filename

        try:
            if config.format == "xlsx":
                # Generate Excel file with charts if requested
                final_path = generate_custom_excel_report(
                    employees, 
                    str(filepath), 
                    config.selectedFields,
                    config.includeCharts,
                    config.name,
                    config.description
                )
            else:
                # Generate CSV file
                df = pd.DataFrame(employees)
                
                # Process list fields for CSV
                for col in df.columns:
                    if df[col].apply(lambda x: isinstance(x, list)).any():
                        df[col] = df[col].apply(lambda x: ", ".join(map(str, x)) if isinstance(x, list) else x)
                
                df.to_csv(filepath, index=False)
                final_path = str(filepath)
                
            # Verify file was created
            if not Path(final_path).exists():
                raise Exception("Report file was not created successfully")
            
            print(f"Generated file path: {final_path}")
            print(f"File exists: {Path(final_path).exists()}")
            print(f"File size: {Path(final_path).stat().st_size if Path(final_path).exists() else 'N/A'}")
                
        except Exception as e:
            raise Exception(f"Report generation failed: {str(e)}")

        # Update database with completion
        custom_reports_collection.update_one(
            {"_id": report_id},
            {"$set": {
                "status": "Completed",
                "file_path": str(final_path),
                "file_name": filename,
                "generated_at": datetime.utcnow().isoformat(),
                "updated_at": datetime.utcnow().isoformat(),
                "message": None
            }}
        )

    except Exception as e:
        # Update status to failed
        custom_reports_collection.update_one(
            {"_id": report_id},
            {"$set": {
                "status": "Failed", 
                "message": str(e),
                "updated_at": datetime.utcnow().isoformat()
            }}
        )

@router.post("/generate")
async def generate_custom_report(config: CustomReportConfig, background_tasks: BackgroundTasks):
    """Generate a custom report based on configuration"""
    try:
        if not config.selectedFields:
            raise HTTPException(status_code=400, detail="At least one field must be selected")

        # Create report record
        report_id = str(uuid.uuid4())
        report_data = {
            "_id": report_id,
            "name": config.name,
            "description": config.description,
            "config": config.dict(),
            "status": "Queued",
            "created_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat()
        }
        
        result = custom_reports_collection.insert_one(report_data)
        if not result.inserted_id:
            raise HTTPException(status_code=500, detail="Failed to create report record")

        # Start background task
        background_tasks.add_task(_generate_custom_report_task, report_id, config)

        return {
            "message": "Custom report generation has been queued successfully.", 
            "report_id": report_id,
            "report_name": config.name
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to queue custom report generation: {str(e)}")

@router.get("/history")
async def get_custom_report_history():
    """Get all custom report generation history"""
    try:
        history = list(custom_reports_collection.find(
            {}, 
            {"_id": 0}
        ).sort("created_at", -1))
        
        return history
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch custom report history: {str(e)}")

@router.get("/{report_id}/download")
async def download_custom_report(report_id: str):
    """Download a completed custom report"""
    try:
        report = custom_reports_collection.find_one({"_id": report_id})
        if not report:
            raise HTTPException(status_code=404, detail="Report not found.")
        
        if report.get("status") != "Completed":
            raise HTTPException(
                status_code=400, 
                detail=f"Report is not ready for download. Current status: {report.get('status', 'Unknown')}"
            )
        
        file_path = Path(report["file_path"])
        if not file_path.exists():
            custom_reports_collection.update_one(
                {"_id": report_id},
                {"$set": {
                    "status": "Failed",
                    "message": "Report file not found on server",
                    "updated_at": datetime.utcnow().isoformat()
                }}
            )
            raise HTTPException(status_code=404, detail="Report file not found on server.")
            
        from fastapi.responses import FileResponse
        
        # Determine media type based on file extension
        media_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" 
            if file_path.suffix.lower() == ".xlsx" 
            else "text/csv"
        )
        
        return FileResponse(
            file_path, 
            filename=report.get("file_name", "custom_report.xlsx"),
            media_type=media_type,
            headers={
                "Content-Disposition": f"attachment; filename=\"{report.get('file_name', 'custom_report.xlsx')}\"",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to download custom report: {str(e)}")
    
    
@router.post("/save")
async def save_custom_report_template(request: SaveTemplateRequest):
    """Save a custom report template for reuse"""
    try:
        template_id = str(uuid.uuid4())
        template_data = {
            "_id": template_id,
            "name": request.name,
            "description": request.description,
            "config": request.config.dict(),
            "created_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat()
        }
        
        result = custom_templates_collection.insert_one(template_data)
        if not result.inserted_id:
            raise HTTPException(status_code=500, detail="Failed to save template")

        return {
            "message": "Report template saved successfully.",
            "template_id": template_id,
            "name": request.name
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save template: {str(e)}")

@router.get("/saved")
async def get_saved_custom_report_templates():
    """Get all saved custom report templates"""
    try:
        templates = list(custom_templates_collection.find({}).sort("created_at", -1))
        
        # Process templates to include proper id field
        processed_templates = []
        for template in templates:
            processed_templates.append({
                "id": str(template["_id"]),
                "name": template["name"],
                "description": template["description"],
                "config": template["config"],
                "created_at": template["created_at"]
            })
        
        return processed_templates
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch saved templates: {str(e)}")

@router.delete("/saved/{template_id}")
async def delete_custom_report_template(template_id: str):
    """Delete a saved custom report template"""
    try:
        result = custom_templates_collection.delete_one({"_id": template_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Template not found.")
        
        return {"message": "Template deleted successfully."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete template: {str(e)}")
    
@router.get("/{report_id}/status")
async def get_custom_report_status(report_id: str):
    """Get status of a specific custom report"""
    try:
        report = custom_reports_collection.find_one({"_id": report_id})
        if not report:
            raise HTTPException(status_code=404, detail="Report not found.")
        
        # Return only the necessary fields for status checking
        return {
            "report_id": report_id,
            "status": report.get("status", "Unknown"),
            "message": report.get("message"),
            "file_name": report.get("file_name"),
            "name": report.get("name"),
            "created_at": report.get("created_at"),
            "updated_at": report.get("updated_at")
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get report status: {str(e)}")


@router.delete("/{report_id}")
async def delete_custom_report(report_id: str):
    """Delete a custom report and its associated file"""
    try:
        report = custom_reports_collection.find_one({"_id": report_id})
        if not report:
            raise HTTPException(status_code=404, detail="Report not found.")
        
        # Delete the physical file if it exists
        file_path = report.get("file_path")
        if file_path and Path(file_path).exists():
            try:
                os.remove(file_path)
            except OSError as e:
                print(f"Warning: Failed to delete custom report file {file_path}: {e}")

        # Delete the document from the database
        result = custom_reports_collection.delete_one({"_id": report_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Report not found in database.")
        
        return {"message": "Custom report and associated file deleted successfully."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete custom report: {str(e)}")

def generate_custom_excel_report(employees, filepath, selected_fields, include_charts, report_name, report_description):
    """Generate a custom Excel report with selected fields and optional charts"""
    if not employees:
        raise ValueError("No employee data provided.")

    df = pd.DataFrame(employees)
    
    # Process list fields for Excel
    for col in df.columns:
        if col in df.columns and df[col].apply(lambda x: isinstance(x, list)).any():
            df[col] = df[col].apply(lambda x: ", ".join(map(str, x)) if isinstance(x, list) else x)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, Reference

    wb = Workbook()
    
    # Main data sheet
    ws = wb.active
    ws.title = "Custom Report Data"
    
    # Add report header
    ws["A1"] = report_name
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = report_description
    ws["A2"].font = Font(italic=True, size=12)
    
    # Add data starting from row 4
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Style the worksheet
    style_custom_worksheet(ws, start_row=4)
    
    # Add summary sheet if charts are requested
    if include_charts and len(employees) > 0:
        summary_ws = wb.create_sheet("Summary & Charts")
        add_custom_charts(summary_ws, df, selected_fields)
    
    wb.save(filepath)
    return filepath

def style_custom_worksheet(ws, start_row=1):
    """Apply styling to custom report worksheet"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    # Style headers
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
        
        # Auto-adjust column widths
        col_letter = get_column_letter(col_num)
        max_length = 15
        for row in ws[col_letter]:
            if row.value:
                max_length = max(max_length, len(str(row.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Style data cells
    for row in ws.iter_rows(min_row=start_row+1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment
            cell.border = thin_border

def add_custom_charts(ws, df, selected_fields):
    """Add charts to summary worksheet based on selected fields"""
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    ws["A1"] = "Report Summary & Analytics"
    ws["A1"].font = Font(bold=True, size=16)
    
    current_row = 3
    
    # Department distribution if department is selected
    if 'department' in selected_fields and 'department' in df.columns:
        dept_summary = df['department'].value_counts().reset_index()
        dept_summary.columns = ['Department', 'Count']
        
        ws[f"A{current_row}"] = "Department Distribution"
        ws[f"A{current_row}"].font = Font(bold=True, size=14)
        current_row += 2
        
        for r_idx, row in enumerate(dataframe_to_rows(dept_summary, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Add chart
        chart = BarChart()
        chart.title = "Employees by Department"
        data = Reference(ws, min_col=2, min_row=current_row, max_row=current_row+len(dept_summary))
        categories = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+len(dept_summary))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, f"D{current_row}")
        
        current_row += len(dept_summary) + 5
    
    # Salary analysis if salary is selected
    if 'salary' in selected_fields and 'salary' in df.columns:
        # Create salary ranges
        salary_bins = [0, 50000, 75000, 100000, 150000, float('inf')]
        salary_labels = ['<$50K', '$50K-$75K', '$75K-$100K', '$100K-$150K', '$150K+']
        df['salary_range'] = pd.cut(df['salary'], bins=salary_bins, labels=salary_labels, right=False)
        salary_summary = df['salary_range'].value_counts().reset_index()
        salary_summary.columns = ['Salary Range', 'Count']
        
        ws[f"A{current_row}"] = "Salary Distribution"
        ws[f"A{current_row}"].font = Font(bold=True, size=14)
        current_row += 2
        
        for r_idx, row in enumerate(dataframe_to_rows(salary_summary, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Add pie chart
        pie_chart = PieChart()
        pie_chart.title = "Salary Distribution"
        data = Reference(ws, min_col=2, min_row=current_row, max_row=current_row+len(salary_summary))
        labels = Reference(ws, min_col=1, min_row=current_row+1, max_row=current_row+len(salary_summary))
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        ws.add_chart(pie_chart, f"D{current_row}")
        
        current_row += len(salary_summary) + 5


