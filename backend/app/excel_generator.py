import os
import uuid
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


# Generate an Excel Report
def generate_employee_report(employees, filename=None):
    if not employees:
        raise ValueError("No employee data provided.")

    os.makedirs("exports", exist_ok=True)

    if not filename:
        filename = f"employee_report_{uuid.uuid4().hex[:6]}.xlsx"

    filepath = os.path.join("exports", filename)

    df = pd.DataFrame(employees)
    df.drop(columns=["_id"], inplace=True, errors='ignore')

    for col in df.columns:
        if df[col].apply(lambda x: isinstance(x, list)).any():
            df[col] = df[col].apply(lambda x: ", ".join(map(str, x)) if isinstance(x, list) else x)

    # Department Summary
    dept_summary = (
        df.groupby("department")
        .agg(employee_count=("employee_id", "count"), avg_salary=("salary", "mean"))
        .reset_index()
    )

    # Salary Range Analysis
    salary_bins = [0, 50_000, 75_000, 100_000, 150_000, 200_000]
    df["salary_range"] = pd.cut(df["salary"], bins=salary_bins)
    
    salary_analysis = (
        df["salary_range"]
        .value_counts()
        .sort_index()
        .reset_index()
    )
    salary_analysis.columns = ["salary_range", "employee_count"]
    salary_analysis["salary_range"] = salary_analysis["salary_range"].astype(str)


    # Excel Workbook Creation
    wb = Workbook()

    # --- Sheet 1: Employee List ---
    ws1 = wb.active
    ws1.title = "Employee List"

    for row in dataframe_to_rows(df.drop(columns=["salary_range"]), index=False, header=True):
        ws1.append(row)

    style_worksheet(ws1)

    # Highlight top performers
    performance_col = df.columns.get_loc("performance_score") + 1
    for row in range(2, ws1.max_row + 1):
        performance = ws1.cell(row=row, column=performance_col).value
        if performance is not None and performance >= 4.5:
            for col in range(1, ws1.max_column + 1):
                ws1.cell(row=row, column=col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # --- Sheet 2: Department Summary ---
    ws2 = wb.create_sheet("Department Summary")

    for row in dataframe_to_rows(dept_summary, index=False, header=True):
        ws2.append(row)

    style_worksheet(ws2)
    title_cell = f"B{ws2.max_row + 2}"
    ws2[title_cell] = "Employees by Department"
    ws2[title_cell].font = Font(bold=True, size=14)

    # Add 2 empty rows after title
    ws2.insert_rows(ws2.max_row + 1, amount=2)

    # Bar Chart - Employees by Department
    chart = BarChart()
    chart.height = 10  # Controls vertical size of the chart

    data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row, max_col=2)
    categories = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws2.add_chart(chart, "B10")  # Positioned lower to avoid title overlap

    # --- Sheet 3: Salary Analysis ---
    ws3 = wb.create_sheet("Salary Analysis")

    for row in dataframe_to_rows(salary_analysis, index=False, header=True):
        ws3.append(row)

    style_worksheet(ws3)
    title_cell = f"B{ws3.max_row + 2}"
    ws3[title_cell] = "Salary Distribution"
    ws3[title_cell].font = Font(bold=True, size=14)

    # Add 2 empty rows after title
    ws3.insert_rows(ws3.max_row + 1, amount=2)

    # Pie Chart - Salary Distribution
    pie = PieChart()

    data = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row)
    labels = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws3.add_chart(pie, "B10")

    # --------------------------------------------
    # Save the Workbook
    # --------------------------------------------
    wb.save(filepath)
    return filepath


def style_worksheet(ws):

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue header
    header_font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Style headers
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

        # Auto-adjust column widths
        col_letter = get_column_letter(col_num)
        max_length = max(len(str(cell.value)) for cell in ws[col_letter])
        ws.column_dimensions[col_letter].width = max(max_length + 2, 15)

    # Style data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment
            cell.border = thin_border


def generate_comprehensive_report(df, filepath):
    """Generate comprehensive employee report with all features"""
    wb = Workbook()
    
    # Sheet 1: Employee List
    ws1 = wb.active
    ws1.title = "Employee Directory"
    
    # Add header and data
    add_report_header(ws1, "Comprehensive Employee Report", "Complete overview of all employees")
    add_dataframe_to_sheet(ws1, df, start_row=4)
    style_worksheet(ws1, start_row=4)
    
    # Sheet 2: Department Analysis
    ws2 = wb.create_sheet("Department Analysis")
    create_department_analysis(ws2, df)
    
    # Sheet 3: Salary Analysis
    ws3 = wb.create_sheet("Salary Analysis")
    create_salary_analysis(ws3, df)
    
    # Sheet 4: Performance Overview
    if 'performance_score' in df.columns:
        ws4 = wb.create_sheet("Performance Overview")
        create_performance_analysis(ws4, df)
    
    # Sheet 5: Summary Dashboard
    ws5 = wb.create_sheet("Executive Summary")
    create_executive_summary(ws5, df)
    
    wb.save(filepath)
    return filepath

def generate_performance_report(df, filepath):
    """Generate performance-focused report"""
    wb = Workbook()
    
    # Filter for performance-related columns
    performance_cols = ['employee_id', 'name', 'department', 'position', 'performance_score', 'join_date']
    available_cols = [col for col in performance_cols if col in df.columns]
    perf_df = df[available_cols] if available_cols else df
    
    # Main performance sheet
    ws1 = wb.active
    ws1.title = "Performance Overview"
    add_report_header(ws1, "Quarterly Performance Report", "Employee performance analysis and trends")
    add_dataframe_to_sheet(ws1, perf_df, start_row=4)
    style_worksheet(ws1, start_row=4)
    
    # Performance analytics
    if 'performance_score' in df.columns:
        ws2 = wb.create_sheet("Performance Analytics")
        create_performance_analysis(ws2, df)
        
        ws3 = wb.create_sheet("Top Performers")
        create_top_performers_analysis(ws3, df)
    
    wb.save(filepath)
    return filepath

def generate_salary_report(df, filepath):
    """Generate salary-focused report"""
    wb = Workbook()
    
    # Filter for salary-related columns
    salary_cols = ['employee_id', 'name', 'department', 'position', 'salary', 'join_date']
    available_cols = [col for col in salary_cols if col in df.columns]
    salary_df = df[available_cols] if available_cols else df
    
    # Main salary sheet
    ws1 = wb.active
    ws1.title = "Salary Overview"
    add_report_header(ws1, "Salary Distribution Report", "Comprehensive salary analysis and trends")
    add_dataframe_to_sheet(ws1, salary_df, start_row=4)
    style_worksheet(ws1, start_row=4)
    
    # Salary analytics
    if 'salary' in df.columns:
        ws2 = wb.create_sheet("Salary Analytics")
        create_salary_analysis(ws2, df)
        
        ws3 = wb.create_sheet("Department Comparison")
        create_department_salary_comparison(ws3, df)
    
    wb.save(filepath)
    return filepath

def generate_department_report(df, filepath):
    """Generate department-focused report"""
    wb = Workbook()
    
    # Main department overview
    ws1 = wb.active
    ws1.title = "Department Overview"
    add_report_header(ws1, "Department Overview Report", "Comprehensive departmental analysis")
    add_dataframe_to_sheet(ws1, df, start_row=4)
    style_worksheet(ws1, start_row=4)
    
    # Department analytics
    if 'department' in df.columns:
        ws2 = wb.create_sheet("Department Analytics")
        create_department_analysis(ws2, df)
        
        # Individual department sheets
        departments = df['department'].unique()
        for dept in departments[:5]:  # Limit to 5 departments to avoid too many sheets
            dept_df = df[df['department'] == dept]
            ws_dept = wb.create_sheet(f"{dept[:20]}")  # Truncate long names
            add_report_header(ws_dept, f"{dept} Department", f"Detailed view of {dept} employees")
            add_dataframe_to_sheet(ws_dept, dept_df, start_row=4)
            style_worksheet(ws_dept, start_row=4)
    
    wb.save(filepath)
    return filepath

def generate_executive_report(df, filepath):
    """Generate executive summary report"""
    wb = Workbook()
    
    # Executive summary sheet
    ws1 = wb.active
    ws1.title = "Executive Summary"
    create_executive_summary(ws1, df)
    
    # Key metrics sheet
    ws2 = wb.create_sheet("Key Metrics")
    create_key_metrics_dashboard(ws2, df)
    
    # Trends analysis
    ws3 = wb.create_sheet("Trends Analysis")
    create_trends_analysis(ws3, df)
    
    wb.save(filepath)
    return filepath

def add_report_header(ws, title, description):
    """Add formatted header to worksheet"""
    ws.merge_cells('A1:E1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A2:E2')
    ws['A2'] = description
    ws['A2'].font = Font(italic=True, size=12)
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Add generation timestamp
    ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'].font = Font(size=10, color="808080")
    

def add_dataframe_to_sheet(ws, df, start_row=1):
    """Add DataFrame data to worksheet starting at specified row"""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Format specific data types
            if isinstance(value, (int, float)) and 'salary' in df.columns:
                salary_col_idx = df.columns.get_loc('salary') + 1 if 'salary' in df.columns else -1
                if c_idx == salary_col_idx:
                    cell.number_format = '$#,##0'
            elif isinstance(value, float) and 'performance_score' in df.columns:
                perf_col_idx = df.columns.get_loc('performance_score') + 1 if 'performance_score' in df.columns else -1
                if c_idx == perf_col_idx:
                    cell.number_format = '0.0'

def create_department_analysis(ws, df):
    """Create department analysis sheet with charts"""
    if 'department' not in df.columns:
        ws['A1'] = "Department data not available"
        return
        
    add_report_header(ws, "Department Analysis", "Employee distribution and metrics by department")
    
    # Department summary
    agg_dict = {'employee_id': 'count'}
    if 'salary' in df.columns:
        agg_dict['salary'] = ['mean', 'median']
    if 'performance_score' in df.columns:
        agg_dict['performance_score'] = 'mean'
    
    dept_summary = df.groupby('department').agg(agg_dict).round(2)
    
    # Flatten column names
    if 'salary' in df.columns and 'performance_score' in df.columns:
        dept_summary.columns = ['Employee_Count', 'Avg_Salary', 'Median_Salary', 'Avg_Performance']
    elif 'salary' in df.columns:
        dept_summary.columns = ['Employee_Count', 'Avg_Salary', 'Median_Salary']
    elif 'performance_score' in df.columns:
        dept_summary.columns = ['Employee_Count', 'Avg_Performance']
    else:
        dept_summary.columns = ['Employee_Count']
        
    dept_summary = dept_summary.reset_index()
    
    # Add data to sheet
    add_dataframe_to_sheet(ws, dept_summary, start_row=5)
    style_worksheet(ws, start_row=5, data_rows=len(dept_summary))
    
    # Add bar chart
    if len(dept_summary) > 1:
        chart = BarChart()
        chart.title = "Employees by Department"
        chart.height = 8
        chart.width = 12
        
        data = Reference(ws, min_col=2, min_row=5, max_row=5+len(dept_summary))
        categories = Reference(ws, min_col=1, min_row=6, max_row=5+len(dept_summary))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "E7")


def create_salary_analysis(ws, df):
    """Create salary analysis sheet with distribution charts"""
    if 'salary' not in df.columns:
        ws['A1'] = "Salary data not available"
        return
        
    add_report_header(ws, "Salary Analysis", "Salary distribution and compensation insights")
    
    # Create salary ranges
    salary_bins = [0, 40000, 60000, 80000, 100000, 120000, float('inf')]
    salary_labels = ['<$40K', '$40K-$60K', '$60K-$80K', '$80K-$100K', '$100K-$120K', '$120K+']
    
    df_copy = df.copy()
    df_copy['salary_range'] = pd.cut(df_copy['salary'], bins=salary_bins, labels=salary_labels, right=False)
    
    salary_analysis = df_copy['salary_range'].value_counts().reindex(salary_labels, fill_value=0).reset_index()
    salary_analysis.columns = ['Salary_Range', 'Employee_Count']
    
    # Add statistics
    salary_stats = pd.DataFrame({
        'Metric': ['Average Salary', 'Median Salary', 'Min Salary', 'Max Salary', 'Std Deviation'],
        'Value': [df['salary'].mean(), df['salary'].median(), df['salary'].min(), df['salary'].max(), df['salary'].std()]
    }).round(2)
    
    # Add data to sheet
    add_dataframe_to_sheet(ws, salary_analysis, start_row=5)
    style_worksheet(ws, start_row=5, data_rows=len(salary_analysis))
    
    # Add statistics table
    start_row_stats = 5 + len(salary_analysis) + 3
    ws[f'A{start_row_stats}'] = "Salary Statistics"
    ws[f'A{start_row_stats}'].font = Font(bold=True, size=14)
    
    add_dataframe_to_sheet(ws, salary_stats, start_row=start_row_stats + 1)
    style_worksheet(ws, start_row=start_row_stats + 1, data_rows=len(salary_stats))
    
    # Add pie chart
    if len(salary_analysis) > 1:
        pie_chart = PieChart()
        pie_chart.title = "Salary Distribution"
        pie_chart.height = 10
        pie_chart.width = 10
        
        data = Reference(ws, min_col=2, min_row=5, max_row=5+len(salary_analysis))
        labels = Reference(ws, min_col=1, min_row=6, max_row=5+len(salary_analysis))
        
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        ws.add_chart(pie_chart, "D7")

def create_performance_analysis(ws, df):
    """Create performance analysis sheet"""
    if 'performance_score' not in df.columns:
        ws['A1'] = "Performance data not available"
        return
        
    add_report_header(ws, "Performance Analysis", "Employee performance metrics and trends")
    
    # Performance ranges
    perf_bins = [0, 2, 3, 4, 4.5, 5]
    perf_labels = ['Poor (0-2)', 'Fair (2-3)', 'Good (3-4)', 'Excellent (4-4.5)', 'Outstanding (4.5-5)']
    
    df_copy = df.copy()
    df_copy['performance_range'] = pd.cut(df_copy['performance_score'], bins=perf_bins, labels=perf_labels, right=False)
    
    perf_analysis = df_copy['performance_range'].value_counts().reindex(perf_labels, fill_value=0).reset_index()
    perf_analysis.columns = ['Performance_Range', 'Employee_Count']
    
    # Performance by department if available
    if 'department' in df.columns:
        perf_by_dept = df.groupby('department')['performance_score'].mean().round(2).reset_index()
        perf_by_dept.columns = ['Department', 'Avg_Performance']
    
    # Add data to sheet
    add_dataframe_to_sheet(ws, perf_analysis, start_row=5)
    style_worksheet(ws, start_row=5, data_rows=len(perf_analysis))
    
    if 'department' in df.columns:
        start_row_dept = 5 + len(perf_analysis) + 3
        ws[f'A{start_row_dept}'] = "Performance by Department"
        ws[f'A{start_row_dept}'].font = Font(bold=True, size=14)
        
        add_dataframe_to_sheet(ws, perf_by_dept, start_row=start_row_dept + 1)
        style_worksheet(ws, start_row=start_row_dept + 1, data_rows=len(perf_by_dept))

def create_top_performers_analysis(ws, df):
    """Create top performers analysis"""
    if 'performance_score' not in df.columns:
        ws['A1'] = "Performance data not available"
        return
        
    add_report_header(ws, "Top Performers", "Highest performing employees")
    
    # Get top performers (score >= 4.5)
    top_performers = df[df['performance_score'] >= 4.5].copy()
    
    if len(top_performers) == 0:
        ws['A5'] = "No employees with performance score >= 4.5 found"
        return
    
    # Select relevant columns
    display_cols = ['employee_id', 'name', 'department', 'position', 'performance_score']
    available_cols = [col for col in display_cols if col in top_performers.columns]
    top_performers_display = top_performers[available_cols].sort_values('performance_score', ascending=False)
    
    add_dataframe_to_sheet(ws, top_performers_display, start_row=5)
    style_worksheet(ws, start_row=5, data_rows=len(top_performers_display))
    
    # Highlight top performers
    for row in range(6, 6 + len(top_performers_display)):
        for col in range(1, len(available_cols) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

def create_executive_summary(ws, df):
    """Create executive summary dashboard"""
    add_report_header(ws, "Executive Summary", "Key metrics and insights at a glance")
    
    # Calculate key metrics
    total_employees = len(df)
    active_employees = len(df[df['is_active'] == True]) if 'is_active' in df.columns else total_employees
    
    metrics = {
        'Total Employees': total_employees,
        'Active Employees': active_employees,
        'Departments': len(df['department'].unique()) if 'department' in df.columns else 'N/A',
        'Average Salary': f"${df['salary'].mean():,.0f}" if 'salary' in df.columns else 'N/A',
        'Average Performance': f"{df['performance_score'].mean():.2f}" if 'performance_score' in df.columns else 'N/A'
    }
    
    # Create metrics summary
    row = 5
    for metric, value in metrics.items():
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Style the metrics section
    for r in range(5, row):
        ws[f'A{r}'].border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                                   top=Side(style="thin"), bottom=Side(style="thin"))
        ws[f'B{r}'].border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                                   top=Side(style="thin"), bottom=Side(style="thin"))

def create_key_metrics_dashboard(ws, df):
    """Create key metrics dashboard"""
    add_report_header(ws, "Key Metrics Dashboard", "Essential business metrics")
    
    # Department distribution
    if 'department' in df.columns:
        ws['A5'] = "Department Distribution"
        ws['A5'].font = Font(bold=True, size=14)
        
        dept_counts = df['department'].value_counts().reset_index()
        dept_counts.columns = ['Department', 'Count']
        
        add_dataframe_to_sheet(ws, dept_counts, start_row=6)
        style_worksheet(ws, start_row=6, data_rows=len(dept_counts))

def create_trends_analysis(ws, df):
    """Create trends analysis sheet"""
    add_report_header(ws, "Trends Analysis", "Historical trends and patterns")
    
    # Hiring trends by join_date if available
    if 'join_date' in df.columns:
        try:
            df_copy = df.copy()
            df_copy['join_date'] = pd.to_datetime(df_copy['join_date'], errors='coerce')
            df_copy['join_year'] = df_copy['join_date'].dt.year
            
            hiring_trends = df_copy['join_year'].value_counts().sort_index().reset_index()
            hiring_trends.columns = ['Year', 'Hires']
            
            ws['A5'] = "Hiring Trends by Year"
            ws['A5'].font = Font(bold=True, size=14)
            
            add_dataframe_to_sheet(ws, hiring_trends, start_row=6)
            style_worksheet(ws, start_row=6, data_rows=len(hiring_trends))
            
        except Exception as e:
            ws['A5'] = f"Unable to parse join_date for trend analysis: {str(e)}"

def create_department_salary_comparison(ws, df):
    """Create department salary comparison"""
    if 'department' not in df.columns or 'salary' not in df.columns:
        ws['A1'] = "Department or salary data not available"
        return
        
    add_report_header(ws, "Department Salary Comparison", "Compensation analysis across departments")
    
    dept_salary = df.groupby('department')['salary'].agg(['mean', 'median', 'min', 'max']).round(0).reset_index()
    dept_salary.columns = ['Department', 'Avg_Salary', 'Median_Salary', 'Min_Salary', 'Max_Salary']
    
    add_dataframe_to_sheet(ws, dept_salary, start_row=5)
    style_worksheet(ws, start_row=5, data_rows=len(dept_salary))
    
    # Format salary columns
    for row in range(6, 6 + len(dept_salary)):
        for col in range(2, 6):  # Salary columns
            cell = ws.cell(row=row, column=col)
            cell.number_format = '$#,##0'

def style_worksheet(ws, start_row=1, data_rows=None):
    """Apply comprehensive styling to worksheet"""
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    # Style headers
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
        
        # Auto-adjust column widths
        col_letter = get_column_letter(col_num)
        max_length = 15
        for row in ws[col_letter]:
            if row.value:
                max_length = max(max_length, len(str(row.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    
    # Style data cells
    end_row = start_row + (data_rows or (ws.max_row - start_row))
    for row in ws.iter_rows(min_row=start_row + 1, max_row=end_row, 
                           min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment
            cell.border = thin_border
            
            # Alternate row colors
            if (cell.row - start_row) % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")




